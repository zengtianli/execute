#!/usr/bin/env python3
import sys
import csv
import os
import glob
from openpyxl import load_workbook


def convert_xlsx_to_csv(input_file, output_file=None, sheet_name=None):
    """
    将xlsx文件转换为csv格式
    :param input_file: 输入的xlsx文件路径
    :param output_file: 输出的csv文件路径，如果不指定则自动生成
    :param sheet_name: 要转换的工作表名称，如果不指定则使用活动工作表
    """
    if output_file is None:
        if sheet_name:
            # 如果指定了工作表，输出文件名包含工作表名
            output_file = f"{input_file.rsplit('.', 1)[0]}_{sheet_name}.csv"
        else:
            output_file = input_file.rsplit('.', 1)[0] + '.csv'

    try:
        wb = load_workbook(input_file, read_only=True, data_only=True)
        
        # 获取要处理的工作表
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                # 将None转换为空字符串，并转换所有非None的单元格值为字符串
                writer.writerow(['' if cell is None else str(cell) for cell in row])

        print(f"转换完成：{input_file} [{sheet_name if sheet_name else '默认工作表'}] -> {output_file}")
        return output_file
    except Exception as e:
        print(f"转换 {input_file} {'工作表 ' + sheet_name if sheet_name else ''} 时出错：{str(e)}")
        return None


def convert_all_sheets_to_csv(input_file):
    """
    将xlsx文件中的所有工作表转换为单独的csv文件
    :param input_file: 输入的xlsx文件路径
    """
    try:
        wb = load_workbook(input_file, read_only=True, data_only=True)
        
        # 获取所有工作表
        sheet_names = wb.sheetnames
        success_count = 0
        
        # 转换每个工作表
        for sheet_name in sheet_names:
            output_file = f"{input_file.rsplit('.', 1)[0]}_{sheet_name}.csv"
            if convert_xlsx_to_csv(input_file, output_file, sheet_name) is not None:
                success_count += 1
        
        # 输出统计信息
        print(f"共处理 {len(sheet_names)} 个工作表，成功转换 {success_count} 个")
        return success_count > 0
    except Exception as e:
        print(f"处理 {input_file} 的多个工作表时出错：{str(e)}")
        return False


def convert_all_xlsx_to_csv(directory=".", recursive=False):
    """
    转换指定目录下的所有xlsx文件为csv
    :param directory: 要处理的目录
    :param recursive: 是否递归处理子目录
    """
    # 确保目录路径以斜杠结尾
    if not directory.endswith(os.sep):
        directory += os.sep

    # 搜索模式
    pattern = "**/*.xlsx" if recursive else "*.xlsx"
    success_count = 0
    total_count = 0
    
    # 查找所有xlsx文件
    for xlsx_file in glob.glob(directory + pattern, recursive=recursive):
        total_count += 1
        if convert_xlsx_to_csv(xlsx_file) is not None:
            success_count += 1
    
    # 输出统计信息
    if total_count == 0:
        print(f"在 {directory} {'及其子目录' if recursive else ''} 中未找到xlsx文件")
    else:
        print(f"共处理 {total_count} 个文件，成功转换 {success_count} 个")


def main():
    import argparse
    
    # 创建命令行参数解析器
    parser = argparse.ArgumentParser(description="将xlsx文件转换为csv格式")
    parser.add_argument("path", nargs="?", default=".", help="要处理的文件或目录路径，默认为当前目录")
    parser.add_argument("-r", "--recursive", action="store_true", help="递归处理子目录")
    parser.add_argument("-o", "--output", help="指定输出文件名（仅对单个文件和单个工作表有效）")
    parser.add_argument("-s", "--sheet", help="指定要转换的工作表名称（仅对单个文件有效）")
    parser.add_argument("-a", "--all-sheets", action="store_true", help="转换所有工作表（仅对单个文件有效）")
    
    args = parser.parse_args()
    
    # 检查路径是文件还是目录
    if os.path.isfile(args.path):
        # 如果是单个文件
        if not args.path.lower().endswith('.xlsx'):
            print("错误：输入文件必须是xlsx格式")
            sys.exit(1)
        try:
            if args.all_sheets:
                # 转换所有工作表
                if not convert_all_sheets_to_csv(args.path):
                    sys.exit(1)
            else:
                # 转换指定工作表或默认工作表
                if convert_xlsx_to_csv(args.path, args.output, args.sheet) is None:
                    sys.exit(1)
        except Exception as e:
            print(f"转换过程中出错：{str(e)}")
            sys.exit(1)
    else:
        # 如果是目录
        convert_all_xlsx_to_csv(args.path, args.recursive)


if __name__ == "__main__":
    main()
